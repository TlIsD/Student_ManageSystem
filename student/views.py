import os
from tkinter.font import names
from django.contrib import auth
from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from openpyxl import load_workbook
from student.models import Student, StudentDetail, Clas, Course, UserInfo


# Create your views here.
def index(request):
    # （1）视图函数单独判断
    # if request.user.is_authenticated:
        # 获取所有的学生数据
        # student_list = Student.objects.all()
        # return render(request, 'student/index.html', {'student_list': student_list})
    # else:
    #     return redirect('/login/')

    # （2）中间件判断
    # 获取所有的学生数据
    student_list = Student.objects.all()
    return render(request, 'student/index.html', {'student_list': student_list})

def add_student(request):

    if request.method == "GET":
        if request.user.is_superuser:
            class_list = Clas.objects.all()
            course_list = Course.objects.all()

            return render(request, 'student/add_student.html', {'class_list': class_list, 'course_list': course_list})
        else:
            err = 1
            return render(request, 'student/error.html', {'err': err})

    else:
        # 获取客户端数据
        print(request.POST)
        # 添加数据到数据库
        # 方式一：
        # name = request.POST.get('name')
        # age = request.POST.get('age')
        # sex = request.POST.get('sex')
        # birthday = request.POST.get('birthday')
        # clas_id = request.POST.get('clas_id')
        # Student.objects.create(name=name, age=age, sex=sex, birthday=birthday, clas_id=clas_id)

        # 方式二：传入键名必须和表中字段名相同（客户端form表单的name值要和models中的字段保持一致）
        stu = Student.objects.create(**request.POST.dict())

        return redirect('/student/')

def delete_student(request, del_id):
    student = Student.objects.get(pk=del_id)

    if request.method == "GET":
        if request.user.is_superuser:
            return render(request, 'student/delete_student.html', {'student': student})
        else:
            err = 1
            return render(request, 'student/error.html', {'err': err})

    else:
        student.delete()
        return redirect('/student/')

def edit_student(request, edit_id):
    student = Student.objects.get(pk=edit_id)
    if request.method == "GET":
        if request.user.is_superuser:
            class_list = Clas.objects.all()
            course_list = Course.objects.all()
            return render(request, 'student/edit_stu.html', {'student': student, 'class_list': class_list, 'course_list': course_list})
        else:
            err = 1
            return render(request, 'student/error.html', {'err': err})
    else:
        print(request.POST)
        # 获取课程id列表
        course_id_list = request.POST.getlist('course_id_list')
        # 获取客户端数据并删除课程id列表
        data = request.POST.dict()
        data.pop('course_id_list')
        # 更新除了多对多以外的数据
        Student.objects.filter(id=edit_id).update(**data)
        # 多对多关系的重制
        student.courses.set(course_id_list)

        return redirect('/student/')

def elective(request):
    course_list = Course.objects.all()
    if request.method == "GET":
        return render(request, 'student/course.html', {'course_list': course_list})
    else:
        print(request.POST)
        course_id_list = request.POST.getlist('course_id_list')

        stu_id = request.user.stu_id

        stu = Student.objects.get(pk=stu_id)
        stu.courses.set(course_id_list)

        return redirect('/student/')

def search(request):
    cate = request.GET.get('cate')
    key_word = request.GET.get('key_word')

    if cate == 'name':
        student_list = Student.objects.filter(name__contains=key_word)
    elif cate == 'class':
        student_list = Student.objects.filter(clas__name__contains=key_word)
    else:
        student_list = []

    return render(request, 'student/index.html', {'student_list': student_list, 'key_word': key_word})

def stu_excel(request):
    if request.user.is_superuser:
        excel_stu = request.FILES.get('excel_stu')
        print(excel_stu)
        print(excel_stu.name)

        # （1）将上传的文件下载到服务器某个文件夹下
        path = os.path.join('media', 'files', excel_stu.name)

        with open(path, 'wb') as f:
            for line in excel_stu:
                f.write(line)

        # （2）通过python操作Excel表格
        wb = load_workbook(path)
        print(wb.sheetnames)
        work_sheet = wb.worksheets[0]
        stu_list = []
        for line in work_sheet.iter_rows(min_row=3):
            # print('line::', line)
            # for cell in line:
            #     print('cell::', cell.value)
            if not line[0].value:
                break

            sd = StudentDetail.objects.create(tel=line[4].value, addr=line[5].value)
            class_id = Clas.objects.get(name=line[6].value).id
            if line[2].value == '男':
                sex = 1
            elif line[2].value == '女':
                sex = 0
            else:
                sex = 2
            stu = Student(name=line[0].value,
                          age=line[1].value,
                          sex=sex,
                          birthday=line[3].value,
                          clas_id=class_id,
                          stu_detail=sd,
                          )
            stu_list.append(stu)

        Student.objects.bulk_create(stu_list)

        return redirect('/student/')

    else:
        err = 1
        return render(request, 'student/error.html', {'err': err})

def login(request):
    if request.method == "GET":
        return render(request, 'student/login.html')
    else:
        user = request.POST.get('loginUsername')
        pwd = request.POST.get('loginPassword')

        # 使用了用户认证组件
        user = auth.authenticate(username=user, password=pwd)
        if user:
            # 验证成功
            # 写session：request.session['user_id']=user.id
            auth.login(request, user)
            return redirect('/')
        else:
            # 失败
            msg = '用户名或密码错误'
            return render(request, 'student/login.html', {'msg': msg})

def logout(request):
    auth.logout(request)
    return redirect('/login/')

def register(request):
    if request.method == "GET":
        return render(request, 'student/register.html')
    else:
        user = request.POST.get('registerUsername')
        stu_id = request.POST.get('registerID')
        pwd = request.POST.get('registerPassword')

        UserInfo.objects.create_user(username=user, password=pwd, stu_id=stu_id)

        msg = '成功'
        print(msg)

        return redirect('/login/')

def judge(request):
    res = {'exist': False, 'user_msg': '', 'id_msg': ''}
    user = request.POST.get('username')
    stu_id = request.POST.get('student_id')

    ret_user = UserInfo.objects.filter(username=user)
    ret_student_id = UserInfo.objects.filter(stu_id=stu_id)

    if ret_user:
        res['exist'] = True
        res['user_msg'] = '该用户已存在'

    if ret_student_id:
        res['exist'] = True
        res['id_msg'] = '该学生已被注册'

    return JsonResponse(res)

def details(request):
    old_username = request.user.username
    if request.method == "GET":

        return render(request, 'student/details.html', {'username': old_username})

    else:
        old_pwd = request.POST.get('old_pwd')
        pwd = request.POST.get('pwd')
        username = request.POST.get('username')
        ret = auth.authenticate(username=username, password=old_pwd)
        print(ret)
        if not ret:
            err = 3
            return render(request, 'student/error.html', {'err': err})

        elif pwd == old_pwd:
            err = 2
            return render(request, 'student/error.html', {'err': err})

        else:
            UserInfo.objects.filter(username=old_username).delete()
            UserInfo.objects.create_user(username=username, password=pwd)
            return redirect('/login/')

def avatar(request):
    if request.method == "POST":
        res = {'status': 'true', 'msg': '上传成功，请刷新'}
        username = request.user.username

        img = request.FILES.get('img')
        print(img)
        suffix = img.name.split('.')[-1]
        print(suffix)

        if suffix == 'jpg' or suffix == 'jpeg' or suffix == 'png':
            # 保存图片并更新数据库
            user_path = os.path.join('media/avatar/',username)
            path = os.path.join('media/avatar/',username, img.name)
            img_path = os.path.join('avatar/', username, img.name)

            if not os.path.exists(user_path):
                os.makedirs(user_path, exist_ok=True)

            with open(path, 'wb') as f:
                for line in img.readlines():
                    f.write(line)

            UserInfo.objects.filter(username=username).update(avatar=img_path)

            # 获取数据
            user = UserInfo.objects.get(username=username)
            print(user.avatar.url)
            res['avatar'] = user.avatar.url

            return JsonResponse(res)

        else:
            res['status'] = 'false'
            res['msg'] = '请上传.jpeg或.jpg或.png格式的文件'
            return JsonResponse(res)